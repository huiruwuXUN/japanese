import openpyxl
import re

def split_characters(input_excel, output_excel):
    try:
        
        wb = openpyxl.load_workbook(input_excel)
        sheet = wb.active
        
        # Create a new workbook for output
        output_wb = openpyxl.Workbook()
        output_sheet = output_wb.active
        
        # Column titles
        output_sheet.append(["RC", "", "", "", "Image Type", "", "Image Number","Hiragana and Katakana","Hiragana and Katakana", "Kanji", "English", "Korean", "Other"])
        
        # Regular expressions to identify character groups
        hiragana_regex = re.compile(r'[\u3040-\u309F]+')  # Hiragana range in Unicode
        katakana_regex = re.compile(r'[\u30A0-\u30FF]+')  # Katakana range in Unicode
        kanji_regex = re.compile(r'[\u4E00-\u9FFF]+')  # Kanji range in Unicode
        english_regex = re.compile(r'[a-zA-Z]+')  # English characters
        korean_regex = re.compile(r'[\uAC00-\uD7AF]+')  # Korean characters
        other_regex = re.compile(r'[^a-zA-Z\u3040-\u309F\u30A0-\u30FF\u4E00-\u9FFF\uAC00-\uD7AF]+')  # Any characters not in the specified sets
        
        # Iterate through each row
        for row in sheet.iter_rows():
            new_row = []
            # Iterate through each cell in the row
            for i, cell in enumerate(row):
                cell_value = cell.value
                if i == 0:
                    # Append the non-character columns (RC, Image Number, Image Type, Image ID)
                    new_row.append(cell_value)
                elif cell_value:
                    # Find character groups
                    hiragana = ''.join(hiragana_regex.findall(cell_value))
                    katakana = ''.join(katakana_regex.findall(cell_value))
                    kanji = ''.join(kanji_regex.findall(cell_value))
                    english = ''.join(english_regex.findall(cell_value))
                    korean = ''.join(korean_regex.findall(cell_value))
                    other = ''.join(other_regex.findall(cell_value))
                    
                    # Append the character groups to the new row
                    new_row.extend([hiragana, katakana, kanji, english, korean, other])
                else:
                    # If cell is empty, append empty strings for each character group
                    new_row.extend([''] * 6)
            
            # Write the new row to the output sheet
            output_sheet.append(new_row)
        
        # Save the output workbook to a new file
        output_wb.save(output_excel)
        print("Splitting characters and exporting to Excel complete.")
    except Exception as e:
        print(f"An error occurred: {e}")

# Example usage:
input_excel_file = "ocr_output.xlsx"  # Change this to your input Excel file
output_excel_file = "split_characters.xlsx"  # Change this to the desired output Excel file
split_characters(input_excel_file, output_excel_file)
