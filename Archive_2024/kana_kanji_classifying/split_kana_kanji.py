from openpyxl import Workbook, load_workbook

def is_kanji(string):
    return all(0x4e00 <= ord(char) <= 0x9fff for char in string)

def is_kana(string):
    return all(0x3040 <= ord(char) <= 0x30ff for char in string)

input_file = "ocr_output.xlsx"
kanji_file = "kanji_output.xlsx"
kana_file = "kana_output.xlsx"
other_file = "other_output.xlsx"

wb_input = load_workbook(filename=input_file)
ws_input = wb_input.active

wb_kanji = Workbook()
ws_kanji = wb_kanji.active
wb_kana = Workbook()
ws_kana = wb_kana.active
wb_other = Workbook()
ws_other = wb_other.active

for row in ws_input.iter_rows(values_only=True):
    if len(row) > 2:
        string = str(row[2])
        if is_kanji(string):
            ws_kanji.append(row)
        elif is_kana(string):
            ws_kana.append(row)
        else:
            ws_other.append(row)

wb_kanji.save(kanji_file)
wb_kana.save(kana_file)
wb_other.save(other_file)